import sys
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import csv
from mpl_toolkits.mplot3d import Axes3D
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn import metrics
from openpyxl import load_workbook
import xlsxwriter
import seaborn as sns

import filter_data as fltr
import models as mdl
 
left   =  0.125  # the left side of the subplots of the figure
right  =  0.9    # the right side of the subplots of the figure
bottom =  0.1    # the bottom of the subplots of the figure
top    =  0.9    # the top of the subplots of the figure
wspace =  0.5     # the amount of width reserved for blank space between subplots
hspace =  1.1    # the amount of height reserved for white space between subplots


#all_cols = ["X", "Y", "Z"]

all_cols = ['acc_x', 'acc_y', 'acc_z', 'gyr_x', 'gyr_y', 'gyr_z', 'mag_x', 'mag_y', 'mag_z', 'type']
features_fltr = ['acc_x', 'acc_y', 'acc_z', 'gyr_x', 'gyr_y', 'gyr_z', 'mag_x', 'mag_y', 'mag_z']


feature_sel = all_cols

float_indexes = features_fltr

f_cols = features_fltr

flt_cols = ['acc_x_blp', 'acc_y_blp', 'acc_z_blp', 'gyr_x_blp', 'gyr_y_blp', 'gyr_z_blp', 'mag_x_blp', 'mag_y_blp', 'mag_z_blp'] 


sns.set(style="darkgrid")

def load_my_data(csvfile, sp):
    dataset = pd.read_csv(csvfile, sep=sp, encoding = "ISO-8859-1", low_memory=False) #encoding='utf-8')
    dataset.fillna(0, inplace=True)
    return dataset

def load_excel_data(xlsfile):
    dataset = pd.read_excel(xlsfile)
    dataset.fillna(0, inplace=True)
    return dataset

    

def convert_to_numeric(dataset, indexes):
    for index in indexes:
        dataset[index] = dataset[index].astype(float)
        #dataset[index] = dataset[index].astype(int)
    return dataset


def select_features(dataset, sel_f):
    X = dataset[sel_f]
    Y = dataset['Label']
    return X,Y


def write_to_file(outdir, inputfile, dset):
        outfile = outdir + "/" + inputfile + ".xlsx"
        create_workbook(outfile, 'filter')
        row = 1
        col = 1
        cl = col
        for i in all_cols:
            lst = dset[i].values.tolist()
            write_to_excel1(outfile, "filter", row, cl , lst, i)
            cl = cl + 1
        for i in flt_cols:
            lst = dset[i].values.tolist()
            write_to_excel1(outfile, "filter", row, cl , lst, i)
            cl = cl + 1  
         
def write_to_excel1(fname, sheetname, row, col, lst, title):
    workbook = load_workbook(filename = fname)
    worksheet = workbook[sheetname]
    worksheet.cell(row=row, column=col).value=title
    for key in range(len(lst)):
        row += 1
        worksheet.cell(row=row, column=col).value = lst[key]
    workbook.save(fname);


def create_workbook(filename, sheetname):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet(sheetname)
    workbook.close()



def preprocess_data(dataset, window, window2):
    dataset = convert_to_numeric(dataset, float_indexes)
    for i in f_cols:
            #dataset = fltr.compute_fn_sliding_window(dataset, i , window)
            #dataset = fltr.apply_kalman(dataset, i)   
            #dataset = fltr.low_pass(dataset, i)  
            #dataset = fltr.band_pass(dataset, i)  
            #dataset = fltr.high_pass(dataset, i)  
            #dataset = fltr.apply_lowess(dataset, i)
            #dataset = fltr.kalman_filter(dataset, i)
            dataset = fltr.butter_low_pass(dataset, i)
    return dataset

def get_accident_probability(user_rating, out_lbls, prob, chances_old):
    nrows = len(out_lbls)
    freq_rash = 0
    for i in range(nrows):
        if prob[i] > 0.4 and out_lbls[i] > 0:
            freq_rash = freq_rash + 1

    chance_acc = freq_rash/nrows
    chances_curr = 0.6 * chance_acc + 0.4 * chances_old        
    rating_curr = user_rating 
    if chances_curr < chances_old:
       rating_curr = (user_rating - (user_rating * 0.1))%10
    else:
       rating_curr = (user_rating + (user_rating * 0.1)) %10
    return chances_curr, rating_curr
    
def main():
    
    if len(sys.argv) <= 2:
        print("Please provide the inputfile(realtimedata)", len(sys.argv))
        exit(-1)
    
    window=20
    window2=10
    user_rating = 5
    chances = 0.2
    inputfile = sys.argv[1]
    out_dir = sys.argv[1].split('.')[0] + "_filtered"
        
    if len(sys.argv) == 3:
        print("User past rating out of 10", int(sys.argv[2]) )
        user_rating = int(sys.argv[2]) 

    baseline_dataset = load_my_data("basedata/main.csv", ",")
    baseline_dataset = preprocess_data(baseline_dataset, window, window2) 
    #print(baseline_dataset)
    m_mdls = mdl.train_model(baseline_dataset) 
    print("DONE TRAINING MODELS WITH BASELINE DATASET")
    #os.mkdir(out_dir)

    filename = inputfile 
    dataset1 = load_my_data(inputfile, ',')
    sliding_samples = 120
    rows = dataset1.shape[0]  
    for timestep in range(int(rows/sliding_samples) * sliding_samples):
       start=timestep
       end=timestep + sliding_samples
       print("PROCESSING WINDOW FROM ", start, " UNTIL ", end) 
       if end > rows:
           end = rows - 1
       if end - start + 1  < 12:
          timestep += sliding_samples
          continue   
       dset1 = dataset1[start:end]
       dset1 = preprocess_data(dset1, window, window2)
       #print(dset1)
       out_lbls, prob = mdl.get_predictions(dset1, m_mdls)
       chances, user_rating = get_accident_probability(user_rating, out_lbls, prob, chances)
       #write_to_file(out_dir, inputfile.split('.')[0], dataset1)
       timestep += sliding_samples
    print("CHANCES OF ACCIDENT:", chances," NEW USER RATING:", round(user_rating))            

if __name__ == "__main__":
    main()
